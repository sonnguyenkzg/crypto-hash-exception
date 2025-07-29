#!/usr/bin/env python3
"""
01_sync_wallet.py - Create Excel file and populate WALLET tab
Usage: python 01_sync_wallet.py
"""

import sys
import os
from pathlib import Path
from datetime import datetime
import pytz
import gspread
from google.oauth2.service_account import Credentials
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List, Dict

# Add src to path
sys.path.insert(0, str(Path(__file__).parent / "src"))

from utils import setup_logging, get_env_variable, get_batch_timestamp_as_datetime

logger = setup_logging()

def get_current_batch() -> str:
    """Read current batch ID from current_batch.txt"""
    try:
        batch_file = Path("current_batch.txt")
        if batch_file.exists():
            batch_id = batch_file.read_text().strip()
            logger.info(f"📄 Found existing batch: {batch_id}")
            return batch_id
        else:
            logger.info("📄 No existing batch file found")
            return None
    except Exception as e:
        logger.warning(f"⚠️ Could not read batch file: {e}")
        return None

def set_current_batch(batch_id: str):
    """Write current batch ID to current_batch.txt"""
    try:
        batch_file = Path("current_batch.txt")
        batch_file.write_text(batch_id)
        logger.info(f"📄 Saved current batch: {batch_id}")
    except Exception as e:
        logger.error(f"❌ Failed to save batch file: {e}")
        raise

def get_current_excel_path(batch_id: str) -> Path:
    """Get Excel file path for current batch"""
    return Path("processed") / f"{batch_id}.xlsx"

class WalletSyncer:
    """Sync wallet data from Google Sheets to Excel"""
    
    def __init__(self):
        self.wallet_sheet_id = get_env_variable('WALLET_SHEET_ID')
        self.wallet_tab_name = get_env_variable('WALLET_SHEET_TAB')
        self.credentials_path = get_env_variable('GOOGLE_SHEETS_CREDENTIALS_PATH', 'config/credentials.json')
        self.client = self._authenticate()
    
    def _authenticate(self):
        """Authenticate with Google Sheets API"""
        scope = [
            'https://spreadsheets.google.com/feeds',
            'https://www.googleapis.com/auth/drive'
        ]
        
        try:
            credentials = Credentials.from_service_account_file(
                self.credentials_path, scopes=scope
            )
            client = gspread.authorize(credentials)
            logger.info("✅ Google Sheets authentication successful")
            return client
            
        except Exception as e:
            logger.error(f"❌ Google Sheets authentication failed: {e}")
            raise
    
    def generate_batch_id(self) -> str:
        """Generate batch ID with GMT+7 timestamp"""
        # Get current time in GMT+7 (Bangkok timezone)
        bangkok_tz = pytz.timezone('Asia/Bangkok')
        now = datetime.now(bangkok_tz)
        
        # Format: YYYYMMDDHHMMSS
        batch_id = now.strftime('%Y%m%d%H%M%S')
        
        logger.info(f"📅 Generated batch ID: {batch_id} (GMT+7: {now.strftime('%Y-%m-%d %H:%M:%S')})")
        return batch_id
    
    def create_processed_folder(self) -> Path:
        """Create processed folder if it doesn't exist"""
        processed_dir = Path("processed")
        processed_dir.mkdir(exist_ok=True)
        logger.info(f"📁 Ensured processed/ folder exists")
        return processed_dir
    
    def read_wallet_data(self, batch_id: str) -> List[Dict]:
        """Read wallet data from Google Sheets"""
        try:
            logger.info(f"📖 Reading wallet data from sheet ID: {self.wallet_sheet_id}")
            logger.info(f"📖 Tab name: {self.wallet_tab_name}")
            
            # Open the workbook and worksheet
            workbook = self.client.open_by_key(self.wallet_sheet_id)
            worksheet = workbook.worksheet(self.wallet_tab_name)
            
            # Get all values
            all_values = worksheet.get_all_values()
            
            if not all_values:
                logger.warning("⚠️ No data found in wallet sheet")
                return []
            
            # Headers should be: Wallet Name, Company, Address, Created At, Refreshed Time
            headers = all_values[0]
            data_rows = all_values[1:]
            
            logger.info(f"📋 Headers: {headers}")
            logger.info(f"📊 Found {len(data_rows)} wallet records")
            
            # Get consistent sync time from batch ID
            sync_time = get_batch_timestamp_as_datetime(batch_id)
            
            wallet_data = []
            
            for row_idx, row in enumerate(data_rows, 2):  # Start from row 2
                if len(row) >= 3 and row[0].strip() and row[2].strip():  # Must have name and address
                    wallet_record = {
                        'wallet_name': row[0].strip() if len(row) > 0 else '',
                        'company': row[1].strip() if len(row) > 1 else '',
                        'address': row[2].strip() if len(row) > 2 else '',
                        'created_at': row[3].strip() if len(row) > 3 else '',
                        'refreshed_time': row[4].strip() if len(row) > 4 else '',
                        'sync_date': sync_time,
                        'source_row': row_idx
                    }
                    
                    # Validate address format (should start with T and be 34 chars)
                    if wallet_record['address'].startswith('T') and len(wallet_record['address']) == 34:
                        wallet_data.append(wallet_record)
                    else:
                        logger.warning(f"⚠️ Invalid address format in row {row_idx}: {wallet_record['address']}")
            
            logger.info(f"✅ Successfully read {len(wallet_data)} valid wallet records")
            return wallet_data
            
        except Exception as e:
            logger.error(f"❌ Failed to read wallet data: {e}")
            raise
    
    def create_excel_file(self, filename: str, processed_dir: Path) -> Path:
        """Create new Excel file"""
        file_path = processed_dir / filename
        
        try:
            # Create new workbook
            workbook = openpyxl.Workbook()
            
            # Remove default sheet
            default_sheet = workbook.active
            workbook.remove(default_sheet)
            
            logger.info(f"📄 Created new Excel file: {file_path}")
            return file_path, workbook
            
        except Exception as e:
            logger.error(f"❌ Failed to create Excel file: {e}")
            raise
    
    def populate_wallet_tab(self, workbook: openpyxl.Workbook, wallet_data: List[Dict]):
        """Populate WALLET tab in Excel"""
        try:
            logger.info(f"📝 Creating WALLET tab with {len(wallet_data)} records...")
            
            # Create WALLET worksheet
            wallet_sheet = workbook.create_sheet("WALLET")
            
            # Define headers - exactly what's in Google Sheets + sync time as last column
            headers = ['Wallet Name', 'Company', 'Address', 'Created At', 'Refreshed Time', 'Sync_Date']
            
            # Style for headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center")
            
            # Write headers
            for col_idx, header in enumerate(headers, 1):
                cell = wallet_sheet.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Write data - exactly as it appears in Google Sheets + sync time from batch
            for row_idx, wallet in enumerate(wallet_data, 2):  # Start from row 2
                wallet_sheet.cell(row=row_idx, column=1, value=wallet['wallet_name'])
                wallet_sheet.cell(row=row_idx, column=2, value=wallet['company'])
                wallet_sheet.cell(row=row_idx, column=3, value=wallet['address'])
                wallet_sheet.cell(row=row_idx, column=4, value=wallet['created_at'])
                wallet_sheet.cell(row=row_idx, column=5, value=wallet['refreshed_time'])
                wallet_sheet.cell(row=row_idx, column=6, value=wallet['sync_date'])
            
            # Auto-adjust column widths
            for column in wallet_sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 chars
                wallet_sheet.column_dimensions[column_letter].width = adjusted_width
            
            logger.info(f"✅ Successfully populated WALLET tab")
            
        except Exception as e:
            logger.error(f"❌ Failed to populate WALLET tab: {e}")
            raise
    
    def save_excel_file(self, workbook: openpyxl.Workbook, file_path: Path):
        """Save Excel file"""
        try:
            workbook.save(file_path)
            logger.info(f"💾 Excel file saved: {file_path}")
            
            # Show file info
            file_size = file_path.stat().st_size
            logger.info(f"📊 File size: {file_size:,} bytes")
            
        except Exception as e:
            logger.error(f"❌ Failed to save Excel file: {e}")
            raise

def main():
    """Main function"""
    try:
        logger.info("🚀 Starting 01_sync_wallet.py")
        logger.info("="*60)
        
        # Initialize syncer
        syncer = WalletSyncer()
        
        # Create processed folder
        processed_dir = syncer.create_processed_folder()
        
        # Generate batch ID and set as current batch
        batch_id = syncer.generate_batch_id()
        set_current_batch(batch_id)
        
        # Create Excel filename
        filename = f"{batch_id}.xlsx"
        excel_path = get_current_excel_path(batch_id)
        
        # Read wallet data from Google Sheets (with batch ID for consistent sync time)
        wallet_data = syncer.read_wallet_data(batch_id)
        
        if not wallet_data:
            logger.error("❌ No wallet data found. Cannot proceed.")
            sys.exit(1)
        
        # Create Excel file
        file_path, workbook = syncer.create_excel_file(filename, processed_dir)
        
        # Populate WALLET tab
        syncer.populate_wallet_tab(workbook, wallet_data)
        
        # Save file
        syncer.save_excel_file(workbook, file_path)
        
        # Final summary
        logger.info("="*60)
        logger.info("🎉 01_sync_wallet.py completed successfully!")
        logger.info(f"📄 Created file: {file_path}")
        logger.info(f"📊 WALLET tab: {len(wallet_data)} records")
        logger.info(f"📅 Batch ID: {batch_id}")
        logger.info(f"📄 Batch file: current_batch.txt")
        
        # Output for pipeline automation
        print(f"BATCH_ID={batch_id}")
        print(f"EXCEL_FILE={file_path}")
        
    except Exception as e:
        logger.error(f"❌ 01_sync_wallet.py failed: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()