#!/usr/bin/env python3
"""
02_sync_ms_form.py - Add MS_FORM tab to existing Excel file
Usage: python 02_sync_ms_form.py
"""

import sys
import os
from pathlib import Path
from datetime import datetime
import pytz
import gspread
from google.oauth2.service_account import Credentials
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List, Dict
import re

# Add src to path
sys.path.insert(0, str(Path(__file__).parent / "src"))

from utils import setup_logging, get_env_variable, get_batch_timestamp_as_datetime

logger = setup_logging()

def get_current_batch() -> str:
    """Read current batch ID from current_batch.txt"""
    try:
        batch_file = Path("current_batch.txt")
        if batch_file.exists():
            batch_id = batch_file.read_text().strip()
            logger.info(f"📄 Current batch: {batch_id}")
            return batch_id
        else:
            logger.error("❌ No current_batch.txt found. Run 01_sync_wallet.py first.")
            sys.exit(1)
    except Exception as e:
        logger.error(f"❌ Could not read batch file: {e}")
        sys.exit(1)

def get_current_excel_path(batch_id: str) -> Path:
    """Get Excel file path for current batch"""
    return Path("processed") / f"{batch_id}.xlsx"

class MSFormSyncer:
    """Sync MS Form data from Google Sheets to Excel"""
    
    def __init__(self):
        self.form_sheet_id = get_env_variable('FORM_SHEET_ID')
        self.form_worksheet = get_env_variable('FORM_WORKSHEET')
        self.credentials_path = get_env_variable('GOOGLE_SHEETS_CREDENTIALS_PATH', 'config/credentials.json')
        self.client = self._authenticate()
        
        # Categories that should result in negative amounts (expenses/reimbursements)
        self.negative_categories = {
            'MARKETING - REIMBURSEMENT',
            'EXPENSE - REIMBURSEMENT', 
            'REFUND',
            'CHARGEBACK',
            'WITHDRAWAL'
        }
    
    def _authenticate(self):
        """Authenticate with Google Sheets API"""
        scope = [
            'https://spreadsheets.google.com/feeds',
            'https://www.googleapis.com/auth/drive'
        ]
        
        try:
            credentials = Credentials.from_service_account_file(
                self.credentials_path, scopes=scope
            )
            client = gspread.authorize(credentials)
            logger.info("✅ Google Sheets authentication successful")
            return client
            
        except Exception as e:
            logger.error(f"❌ Google Sheets authentication failed: {e}")
            raise
    
    def read_form_data(self, batch_id: str) -> List[Dict]:
        """Read form data from Google Sheets"""
        try:
            logger.info(f"📖 Reading form data from sheet ID: {self.form_sheet_id}")
            logger.info(f"📖 Worksheet: {self.form_worksheet}")
            
            # Open the workbook and worksheet
            workbook = self.client.open_by_key(self.form_sheet_id)
            worksheet = workbook.worksheet(self.form_worksheet)
            
            # Get all values
            all_values = worksheet.get_all_values()
            
            if not all_values:
                logger.warning("⚠️ No data found in form sheet")
                return []
            
            headers = all_values[0]
            data_rows = all_values[1:]
            
            logger.info(f"📋 Form headers ({len(headers)} columns): {headers}")
            logger.info(f"📊 Found {len(data_rows)} rows in form data")
            
            # Get consistent sync time from batch ID
            sync_time = get_batch_timestamp_as_datetime(batch_id)
            
            form_data = []
            skipped_empty_hash = 0
            
            for row_idx, row in enumerate(data_rows, 2):  # Start from row 2
                # Skip empty rows
                if not any(cell.strip() for cell in row if cell):
                    continue
                
                # Find hash column and check if it has value
                hash_col_idx = None
                hash_value = None
                
                for col_idx, header in enumerate(headers):
                    if 'hash' in header.lower():
                        hash_col_idx = col_idx
                        hash_value = row[col_idx].strip() if col_idx < len(row) else ''
                        break
                
                # Skip rows with empty hash column
                if hash_col_idx is not None and not hash_value:
                    skipped_empty_hash += 1
                    logger.debug(f"Skipping row {row_idx}: empty hash column")
                    continue
                
                # Create record with all columns from source
                record = {'form_row': row_idx, 'sync_date': sync_time}
                
                # Add all columns dynamically using headers
                for col_idx, header in enumerate(headers):
                    cell_value = row[col_idx].strip() if col_idx < len(row) else ''
                    
                    # Clean header name for dict key (remove spaces, special chars)
                    clean_header = re.sub(r'[^a-zA-Z0-9_]', '_', header).strip('_').lower()
                    if not clean_header:
                        clean_header = f'column_{col_idx}'
                    
                    record[clean_header] = cell_value
                
                # Apply special processing for key fields if they exist
                if hash_value:  # Only process if hash exists
                    record['clean_txn_hash'] = self._extract_hash_from_url(hash_value)
                
                # Apply amount processing if amount column exists
                amount_key = next((k for k in record.keys() if 'amount' in k.lower() and 'expense' not in k.lower()), None)
                category_key = next((k for k in record.keys() if 'category' in k.lower()), None)
                
                if amount_key and record[amount_key]:
                    category = record[category_key].upper() if category_key and record[category_key] else ''
                    record['processed_amount'] = self._extract_amount_with_category_logic(record[amount_key], category)
                
                form_data.append(record)
                
            if skipped_empty_hash > 0:
                logger.info(f"⏭️ Skipped {skipped_empty_hash} rows with empty hash columns")
                
            logger.info(f"✅ Processed {len(form_data)} form records with all {len(headers)} columns")
            
            return form_data
            
        except Exception as e:
            logger.error(f"❌ Failed to read form data: {e}")
            raise
    
    def _extract_hash_from_url(self, raw_hash: str) -> str:
        """Extract transaction hash from URL or return as-is"""
        if not raw_hash:
            return ''
        
        # Check if it's a TronScan URL
        if 'tronscan.org' in raw_hash:
            # Extract hash from URL like: https://tronscan.org/#/transaction/HASH_HERE
            match = re.search(r'/transaction/([a-fA-F0-9]{64})', raw_hash)
            if match:
                return match.group(1)
        
        # If it looks like a direct hash (64 hex characters)
        if re.match(r'^[a-fA-F0-9]{64}$', raw_hash):
            return raw_hash
        
        # Return as-is if we can't parse it
        return raw_hash
    
    def _extract_amount_with_category_logic(self, amount_str: str, category: str) -> float:
        """Extract numeric amount from string and apply category-based sign logic"""
        if not amount_str:
            return 0.0
        
        try:
            # Remove commas, spaces, dollar signs, but preserve parentheses for now
            cleaned = re.sub(r'[,$\s]', '', amount_str)
            
            # Check if amount is in parentheses (accounting format for negative)
            is_parentheses_negative = cleaned.startswith('(') and cleaned.endswith(')')
            if is_parentheses_negative:
                cleaned = cleaned.strip('()')
            
            # Convert to float
            base_amount = float(cleaned) if cleaned else 0.0
            
            # If amount was in parentheses, it's negative
            if is_parentheses_negative:
                base_amount = -abs(base_amount)
            else:
                # Apply category-based logic
                if category in self.negative_categories:
                    # Categories like MARKETING - REIMBURSEMENT should be positive
                    base_amount = abs(base_amount)
                else:
                    # Default categories should be negative
                    base_amount = -abs(base_amount)
            
            return base_amount
            
        except Exception as e:
            logger.warning(f"⚠️ Could not parse amount: '{amount_str}' for category '{category}': {e}")
            return 0.0
    
    def add_ms_form_tab(self, excel_path: Path, form_data: List[Dict]):
        """Add MS_FORM tab to existing Excel file"""
        try:
            logger.info(f"📝 Adding MS_FORM tab to {excel_path}")
            
            # Open existing Excel file
            workbook = openpyxl.load_workbook(excel_path)
            
            # Remove MS_FORM sheet if it already exists
            if "MS_FORM" in workbook.sheetnames:
                workbook.remove(workbook["MS_FORM"])
                logger.info("🗑️ Removed existing MS_FORM tab")
            
            # Create MS_FORM worksheet
            ms_form_sheet = workbook.create_sheet("MS_FORM")
            
            if not form_data:
                logger.warning("⚠️ No data to write to MS_FORM tab")
                return
            
            # Get all unique column names from all records
            all_columns = set()
            for record in form_data:
                all_columns.update(record.keys())
            
            # Sort columns for consistent order, put important ones first
            priority_columns = ['form_row', 'clean_txn_hash', 'processed_amount', 'sync_date']
            sorted_columns = []
            
            # Add priority columns first (if they exist)
            for col in priority_columns:
                if col in all_columns:
                    sorted_columns.append(col)
                    all_columns.remove(col)
            
            # Add remaining columns alphabetically
            sorted_columns.extend(sorted(all_columns))
            
            logger.info(f"📊 Writing {len(sorted_columns)} columns: {sorted_columns[:10]}{'...' if len(sorted_columns) > 10 else ''}")
            
            # Style for headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            header_alignment = Alignment(horizontal="center")
            
            # Write headers
            for col_idx, column_name in enumerate(sorted_columns, 1):
                cell = ms_form_sheet.cell(row=1, column=col_idx, value=column_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Write data
            for row_idx, record in enumerate(form_data, 2):  # Start from row 2
                for col_idx, column_name in enumerate(sorted_columns, 1):
                    value = record.get(column_name, '')
                    cell = ms_form_sheet.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Special formatting for processed amounts
                    if column_name == 'processed_amount' and isinstance(value, (int, float)) and value < 0:
                        cell.font = Font(color="FF0000")  # Red for negative amounts
            
            # Auto-adjust column widths
            for column in ms_form_sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 60)  # Cap at 60 chars
                ms_form_sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save the workbook
            workbook.save(excel_path)
            
            logger.info(f"✅ Successfully added MS_FORM tab with {len(form_data)} records and {len(sorted_columns)} columns")
            
            # Show summary if processed amounts exist
            if any('processed_amount' in record for record in form_data):
                amounts = [record.get('processed_amount', 0) for record in form_data if isinstance(record.get('processed_amount'), (int, float))]
                if amounts:
                    total_amount = sum(amounts)
                    positive_amount = sum(amt for amt in amounts if amt > 0)
                    negative_amount = sum(amt for amt in amounts if amt < 0)
                    
                    logger.info(f"💰 Total net amount: ${total_amount:,.2f}")
                    logger.info(f"📈 Positive amounts: ${positive_amount:,.2f}")
                    logger.info(f"📉 Negative amounts: ${negative_amount:,.2f}")
            
        except Exception as e:
            logger.error(f"❌ Failed to add MS_FORM tab: {e}")
            raise

def main():
    """Main function"""
    try:
        logger.info("🚀 Starting 02_sync_ms_form.py")
        logger.info("="*60)
        
        # Get current batch
        batch_id = get_current_batch()
        excel_path = get_current_excel_path(batch_id)
        
        # Check if Excel file exists
        if not excel_path.exists():
            logger.error(f"❌ Excel file not found: {excel_path}")
            logger.error("Run 01_sync_wallet.py first to create the Excel file")
            sys.exit(1)
        
        logger.info(f"📄 Working with batch: {batch_id}")
        logger.info(f"📄 Excel file: {excel_path}")
        
        # Initialize syncer
        syncer = MSFormSyncer()
        
        # Read form data from Google Sheets (with batch ID for consistent sync time)
        form_data = syncer.read_form_data(batch_id)
        
        if not form_data:
            logger.warning("⚠️ No valid form data found")
            logger.info("This could mean:")
            logger.info("  1. No form submissions with valid hash + amount")
            logger.info("  2. Form sheet is empty")
            logger.info("  3. Column mappings are incorrect")
            sys.exit(1)
        
        # Add MS_FORM tab to Excel file
        syncer.add_ms_form_tab(excel_path, form_data)
        
        # Final summary
        logger.info("="*60)
        logger.info("🎉 02_sync_ms_form.py completed successfully!")
        logger.info(f"📄 Updated file: {excel_path}")
        logger.info(f"📊 MS_FORM tab: {len(form_data)} records")
        logger.info(f"📅 Batch ID: {batch_id}")
        
        # Output for pipeline automation
        print(f"BATCH_ID={batch_id}")
        print(f"EXCEL_FILE={excel_path}")
        print(f"MS_FORM_RECORDS={len(form_data)}")
        
    except Exception as e:
        logger.error(f"❌ 02_sync_ms_form.py failed: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()